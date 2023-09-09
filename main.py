import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from datetime import datetime
import openpyxl
from openpyxl import Workbook

# Function to clear the input fields and update the date
def refresh_form():
    name_entry.delete(0, tk.END)
    class_var.set("FYCS")
    roll_number_entry.delete(0, tk.END)
    attendance_var.set("Present")
    
    # Get today's date and set it as the default value for the date entry
    today_date = datetime.today().strftime("%Y-%m-%d")
    date_picker.delete(0, tk.END)
    date_picker.insert(0, today_date)

# Function to submit the form
def submit_form():
    name = name_entry.get()
    student_class = class_var.get()
    roll_number = roll_number_entry.get()
    attendance_status = attendance_var.get()
    date_str = date_picker.get()
    
    # Convert the selected date to a numeric format (YYYYMMDD)
    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        date_numeric = int(date_obj.strftime("%Y%m%d"))
    except ValueError:
        messagebox.showerror("Error", "Invalid date format. Please use YYYY-MM-DD.")
        return
    
    # You can add further processing or storage of the data here
    # For now, we'll just display it in a messagebox
    message = f"Name: {name}\nClass: {student_class}\nRoll Number: {roll_number}\nAttendance: {attendance_status}\nDate (Numeric): {date_numeric}"
    messagebox.showinfo("Form Submitted", message)

    #Submission of form 
    name = name_entry.get()
    student_class = class_var.get()
    roll_number = roll_number_entry.get()
    attendance_status = attendance_var.get()
    date_str = date_picker.get()

    # Convert the selected date to a numeric format (YYYYMMDD)
    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        date_numeric = int(date_obj.strftime("%Y%m%d"))
    except ValueError:
        messagebox.showerror("Error", "Invalid date format. Please use YYYY-MM-DD.")
        return

    # Create or open the Excel workbook
    excel_filename = "report.xlsx"
    try:
        workbook = openpyxl.load_workbook(excel_filename)
    except FileNotFoundError:
        workbook = Workbook()

    # Select the active sheet or create a new one
    sheet = workbook.active

    # If the sheet doesn't have headers, add them
    if sheet.max_row == 0:
        sheet.append(["Name", "Class", "Roll Number", "Attendance", "Date (Numeric)"])

    # Add data to the Excel sheet
    sheet.append([name, student_class, roll_number, attendance_status, today_date])

    # Save the Excel file
    workbook.save(excel_filename)

    messagebox.showinfo("Form Submitted", "Data saved to report.xlsx")
    
    # Refresh the form
    refresh_form()

# Create the main window
root = tk.Tk()
root.title("Student Information Form")

# Increase the window size
root.geometry("400x300")

# Create a label for the name
name_label = tk.Label(root, text="Name:")
name_label.pack()

# Create a textbox for the name
name_entry = tk.Entry(root)
name_entry.pack()

# Create a label for the class
class_label = tk.Label(root, text="Class:")
class_label.pack()

# Create radiobuttons for class selection
class_var = tk.StringVar()
class_var.set("FYCS")  # Default selection
fy_radio = ttk.Radiobutton(root, text="FYCS", variable=class_var, value="FYCS")
sy_radio = ttk.Radiobutton(root, text="SYCS", variable=class_var, value="SYCS")
ty_radio = ttk.Radiobutton(root, text="TYCS", variable=class_var, value="TYCS")
fy_radio.pack()
sy_radio.pack()
ty_radio.pack()

# Create a label for the roll number
roll_number_label = tk.Label(root, text="Roll Number:")
roll_number_label.pack()

# Create a number entry for the roll number
roll_number_entry = tk.Entry(root)
roll_number_entry.pack()

# Create a label for attendance
attendance_label = tk.Label(root, text="Attendance:")
attendance_label.pack()

# Create radiobuttons for attendance status
attendance_var = tk.StringVar()
attendance_var.set("Present")  # Default selection
present_radio = ttk.Radiobutton(root, text="Present", variable=attendance_var, value="Present")
absent_radio = ttk.Radiobutton(root, text="Absent", variable=attendance_var, value="Absent")
present_radio.pack()
absent_radio.pack()

# Get today's date and set it as the default value for the date entry
today_date = datetime.today().strftime("%Y-%m-%d")
date_label = tk.Label(root, text="Date (YYYY-MM-DD):")
date_label.pack()
date_picker = tk.Entry(root)
date_picker.insert(0, today_date)
date_picker.pack()

# Create a button to submit the form
submit_button = tk.Button(root, text="Submit", command=submit_form)
submit_button.pack()

# # Create a button to refresh the form
# refresh_button = tk.Button(root, text="Refresh", command=refresh_form)
# refresh_button.pack()

# Start the Tkinter main loop
root.mainloop()

# Function to submit the form
def submit_form():
    name = name_entry.get()
    student_class = class_var.get()
    roll_number = roll_number_entry.get()
    attendance_status = attendance_var.get()
    date_str = date_picker.get()

    # Convert the selected date to a numeric format (YYYYMMDD)
    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        date_numeric = int(date_obj.strftime("%Y%m%d"))
    except ValueError:
        messagebox.showerror("Error", "Invalid date format. Please use YYYY-MM-DD.")
        return

    # Create or open the Excel workbook
    excel_filename = "report.xlsx"
    try:
        workbook = openpyxl.load_workbook(excel_filename)
    except FileNotFoundError:
        workbook = Workbook()

    # Select the active sheet or create a new one
    sheet = workbook.active

    # If the sheet doesn't have headers, add them
    if sheet.max_row == 0:
        sheet.append(["Name", "Class", "Roll Number", "Attendance", "Date (Numeric)"])

    # Add data to the Excel sheet
    sheet.append([name, student_class, roll_number, attendance_status, date_numeric])

    # Save the Excel file
    workbook.save(excel_filename)

    messagebox.showinfo("Form Submitted", "Data saved to report.xlsx")